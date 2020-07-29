import openpyxl, random

wb = openpyxl.load_workbook('10주차.xlsx')
sheet = wb['Sheet1']
row_range = sheet[1:50]

#Past tense
def pasttense() :
    for verb in range(1,51):
        sub = sheet['A{}'.format(random.randint(1,6))]
        sheet.cell(row = verb, column = 6, value=sub.value + sheet['B{}'.format(verb)].value)

#Past continuous
def pastcontinuous() :
    for verb in range(1,51):
        sub = sheet['A{}'.format(random.randint(1,6))]
        if sub.row <= 3 :
            sheet.cell(row = verb, column = 7, value=sub.value + 'was ' + sheet['C{}'.format(verb)].value)
        else :
            sheet.cell(row = verb, column = 7, value=sub.value + 'were ' + sheet['C{}'.format(verb)].value)

#Past perfect
def pastperfect() :
    for verb in range(1,51):
        sub = sheet['A{}'.format(random.randint(1,6))]
        sheet.cell(row = verb, column = 8, value=sub.value + 'had ' + sheet['D{}'.format(verb)].value)

#Past perfect continuous
def pastperfectcontinuous() :
    for verb in range(1,51):
        sub = sheet['A{}'.format(random.randint(1,6))]
        sheet.cell(row = verb, column = 9, value=sub.value + 'had been ' + sheet['D{}'.format(verb)].value)



#Present continuous
def presentcontinuous() :
    for verb in range(1,51):
        sub = sheet['A{}'.format(random.randint(1,6))]
        if sub.row == 1 :
            sheet.cell(row = verb, column = 11, value=sub.value + 'am ' + sheet['C{}'.format(verb)].value)
        elif sub.row <= 3 :
            sheet.cell(row = verb, column = 11, value=sub.value + 'is ' + sheet['C{}'.format(verb)].value)
        else :
            sheet.cell(row = verb, column = 11, value=sub.value + 'are ' + sheet['C{}'.format(verb)].value)

#Present perfect
def presentperfect() :
    for verb in range(1,51):
        sub = sheet['A{}'.format(random.randint(1,6))]
        if (sub.row != 1) and (sub.row <= 3) :
            sheet.cell(row = verb, column = 12, value=sub.value + 'has ' + sheet['C{}'.format(verb)].value)
        else :
            sheet.cell(row = verb, column = 12, value=sub.value + 'have ' + sheet['C{}'.format(verb)].value)



#Present perfect continuous
def presentperfectcontinuous() :
    for verb in range(1,51):
        sub = sheet['A{}'.format(random.randint(1,6))]
        if (sub.row != 1) and (sub.row <= 3) :
            sheet.cell(row = verb, column = 13, value=sub.value + 'has been ' + sheet['C{}'.format(verb)].value)
        else :
            sheet.cell(row = verb, column = 13, value=sub.value + 'have been ' + sheet['C{}'.format(verb)].value)

#Future tense
def futuresimple() :
    for verb in range(1,51):
        sub = sheet['A{}'.format(random.randint(1,6))]
        sheet.cell(row = verb, column = 18, value=sub.value + 'will ' + sheet['E{}'.format(verb)].value)

#Future continous
def futurecontinuous() :
    for verb in range(1,51):
        sub = sheet['A{}'.format(random.randint(1,6))]
        sheet.cell(row = verb, column = 15, value=sub.value + 'will be '+ sheet['C{}'.format(verb)].value)

#Future perfect
def futureperfect() :
    for verb in range(1,51):
        sub = sheet['A{}'.format(random.randint(1,6))]
        sheet.cell(row = verb, column = 16, value=sub.value + 'will have ' + sheet['D{}'.format(verb)].value)

#Future perfect continuous
def futureperfectcontinuous() :
    for verb in range(1,51):
        sub = sheet['A{}'.format(random.randint(1,6))]
        sheet.cell(row = verb, column = 17, value=sub.value + 'will have been ' + sheet['C{}'.format(verb)].value)

pasttense()
pastcontinuous()
pastperfect()
pastperfectcontinuous()
presentcontinuous()
presentperfect()
presentperfectcontinuous()
futuresimple()
futurecontinuous()
futureperfect()
futureperfectcontinuous()
wb.save('10주차1.xlsx')
